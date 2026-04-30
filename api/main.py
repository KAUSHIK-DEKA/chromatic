"""
FastAPI backend for Chromatic.

Endpoints:
  GET  /                       -> serve frontend
  GET  /health                 -> {"status": "ok"}

  Batch:
  POST /extract                -> zip in, xlsx out (Base + Printed columns)

  Single-image + calibration:
  POST /classify-single        -> one image (multipart OR base64). Returns
                                  full colors list with roles.
  POST /correct                -> submit a correction for one color (role).
  GET  /corrections/export     -> download corrections.json
  POST /corrections/import     -> upload corrections.json (merge|replace)
  GET  /corrections/stats      -> counts
  POST /corrections/reset      -> wipe all
"""
import base64
import io
import json
import os
import re
import zipfile
from pathlib import Path
from typing import Literal, Optional

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, Field

from color_extractor import (
    add_correction,
    get_corrections_snapshot,
    get_corrections_stats,
    hash_image_bytes,
    import_corrections,
    process_image,
    reset_corrections,
)

app = FastAPI(title="Chromatic — Color Extractor")

STATIC_DIR = Path(__file__).parent.parent / "static"
IMAGE_EXTS = {".webp", ".jpg", ".jpeg", ".png"}
MAX_SINGLE_IMAGE_BYTES = 10 * 1024 * 1024


# ---------------- Frontend ----------------
@app.get("/")
def serve_frontend():
    html_path = STATIC_DIR / "index.html"
    if html_path.exists():
        return FileResponse(html_path)
    raise HTTPException(status_code=404, detail="index.html not found")


if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


@app.get("/health")
def health():
    return {"status": "ok"}


# ---------------- Batch extract ----------------
def _row_for_result(sku: str, result: dict):
    """Build the xlsx row from a process_image result.

    Schema:
      SKU | Base Color | Base Color Code | Base Parent
          | Printed Color | Printed Color Code | Printed Parent
          | Multi
    """
    colors = result.get("colors", [])
    is_multi = result.get("is_multi", False)

    # Base column
    if is_multi:
        base_name = "Multi"
        base_hex = ""
        base_parent = "Multi"
    else:
        if colors:
            base_name = colors[0]["name"]
            base_hex = colors[0]["hex"]
            base_parent = colors[0]["parent"]
        else:
            base_name = base_hex = base_parent = ""

    # Printed column — second-most-dominant color, if any
    if len(colors) >= 2:
        print_name = colors[1]["name"]
        print_hex = colors[1]["hex"]
        print_parent = colors[1]["parent"]
    else:
        print_name = print_hex = print_parent = ""

    multi_marker = "Multi" if is_multi else ""

    return (sku, base_name, base_hex, base_parent, print_name, print_hex, print_parent, multi_marker)


@app.post("/extract")
async def extract_colors_endpoint(
    file: UploadFile = File(...),
    mode: Literal["garment", "footwear"] = Form("garment"),
):
    if not file.filename.lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="Please upload a .zip file")

    zip_bytes = await file.read()
    if len(zip_bytes) == 0:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    rows = []
    errors = []

    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
            entries = []
            for name in zf.namelist():
                if name.endswith("/"):
                    continue
                basename = os.path.basename(name)
                if not basename or basename.startswith("."):
                    continue
                if "__MACOSX" in name:
                    continue
                ext = os.path.splitext(basename)[1].lower()
                if ext in IMAGE_EXTS:
                    entries.append((basename, name))

            if not entries:
                raise HTTPException(
                    status_code=400,
                    detail="No image files (.webp/.jpg/.jpeg/.png) found in the zip",
                )

            entries.sort(key=lambda x: x[0])

            for basename, zip_path in entries:
                sku = os.path.splitext(basename)[0]
                try:
                    with zf.open(zip_path) as img_file:
                        raw = img_file.read()
                    result = process_image(io.BytesIO(raw), mode=mode, image_bytes=raw)
                    rows.append(_row_for_result(sku, result))
                except Exception as e:
                    rows.append((sku, "ERROR", "", str(e)[:60], "", "", "", ""))
                    errors.append(f"{sku}: {e}")
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="Invalid or corrupted zip file")

    # Build xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = mode.capitalize()

    headers = [
        "SKU",
        "Base Color", "Base Color Code", "Base Parent",
        "Printed Color", "Printed Color Code", "Printed Parent",
        "Multi",
    ]
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="2F5597")
    header_align = Alignment(horizontal="center", vertical="center")
    body_font = Font(name="Arial", size=11)
    body_align = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    for i, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = border

    # Column widths
    widths = [18, 22, 16, 18, 22, 16, 18, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    source_name = os.path.splitext(file.filename)[0]
    out_name = f"{source_name}_colors.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{out_name}"',
            "X-Processed-Count": str(len(rows)),
            "X-Error-Count": str(len(errors)),
        },
    )


# ---------------- Single-image classify ----------------
@app.post("/classify-single")
async def classify_single(
    file: Optional[UploadFile] = File(None),
    image_b64: Optional[str] = Form(None),
    mode: Literal["garment", "footwear"] = Form("garment"),
):
    if file is None and not image_b64:
        raise HTTPException(status_code=400, detail="Provide either a file or image_b64")

    if file is not None:
        raw = await file.read()
    else:
        b64_clean = image_b64.split(",", 1)[1] if "," in image_b64 else image_b64
        try:
            raw = base64.b64decode(b64_clean)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid base64 image data")

    if not raw:
        raise HTTPException(status_code=400, detail="Empty image data")
    if len(raw) > MAX_SINGLE_IMAGE_BYTES:
        raise HTTPException(status_code=400, detail="Image too large (max 10MB)")

    try:
        result = process_image(io.BytesIO(raw), mode=mode, image_bytes=raw)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to process image: {e}")

    return JSONResponse(result)


# ---------------- Corrections ----------------
class CorrectionRequest(BaseModel):
    image_hash: Optional[str] = None
    sampled_rgb: list = Field(..., min_length=3, max_length=3)
    corrected_hex: str
    corrected_name: Optional[str] = None
    corrected_family: Optional[str] = None
    role: Optional[str] = None  # 'base' | 'print' | 'print2' (optional)
    # Full colors list (with the corrected role substituted) for pinning the
    # whole image. Omit to skip image-hash override.
    full_colors_for_pin: Optional[list] = None


_HEX_RE = re.compile(r"^#?[0-9A-Fa-f]{6}$")


@app.post("/correct")
async def correct(req: CorrectionRequest):
    if not _HEX_RE.match(req.corrected_hex):
        raise HTTPException(status_code=400, detail="corrected_hex must be a 6-digit hex color")
    try:
        sampled_rgb = [int(c) for c in req.sampled_rgb]
        if any(c < 0 or c > 255 for c in sampled_rgb):
            raise ValueError("rgb out of range")
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="sampled_rgb must be 3 ints 0-255")

    name = (req.corrected_name or "").strip() or None
    family = (req.corrected_family or "").strip() or None

    saved = add_correction(
        image_hash=req.image_hash,
        sampled_rgb=sampled_rgb,
        corrected_hex=req.corrected_hex,
        corrected_name=name,
        corrected_family=family,
        role=req.role,
        full_colors_for_pin=req.full_colors_for_pin,
    )
    return {"ok": True, "saved": saved, "stats": get_corrections_stats()}


@app.get("/corrections/export")
def export_corrections():
    snapshot = get_corrections_snapshot()
    buf = io.BytesIO(json.dumps(snapshot, indent=2).encode("utf-8"))
    return StreamingResponse(
        buf,
        media_type="application/json",
        headers={"Content-Disposition": 'attachment; filename="corrections.json"'},
    )


@app.post("/corrections/import")
async def import_corrections_endpoint(
    file: UploadFile = File(...),
    mode: Literal["merge", "replace"] = Form("merge"),
):
    raw = await file.read()
    try:
        payload = json.loads(raw)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON: {e}")
    stats = import_corrections(payload, mode=mode)
    return {"ok": True, "stats": stats}


@app.get("/corrections/stats")
def corrections_stats():
    return get_corrections_stats()


@app.post("/corrections/reset")
def corrections_reset():
    return reset_corrections()


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.environ.get("PORT", 8000)))
